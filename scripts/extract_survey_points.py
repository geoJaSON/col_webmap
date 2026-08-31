"""
Convert the TPWD ground-sample workbook -> data/survey_points.json + supabase/survey_seed.sql

TPWD assigns every ground sample: a coordinate (decimal degrees, WGS84) and a
type, on-reef or off-reef, which decides which datasheet the crew fills in.
This reads that assignment and emits both a JSON copy for the app and
idempotent SQL for Supabase.

One worksheet per application, named "<app#> (<TPWD site code>)". Point numbers
restart at 1 in each sheet, so the key is (app_no, point_no) -- not point_no
alone.

Layout note: columns A-F are the per-point rows. Columns H-I are a two-row
summary of on-reef / off-reef *acreage* for the site, parked beside the data
rather than below it. They are not per-point values and are read separately;
the per-point classification is the `Reef` column (E).
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "survey_points" / "Woody_Jurisich_GB_NRS_1.xlsx"
JSON_OUT = ROOT / "data" / "survey_points.json"
SQL_OUT = ROOT / "supabase" / "survey_seed.sql"

# The workbook writes these with inconsistent spacing ("On Reef", "on  reef").
REEF_TYPES = {"on reef": "on", "off reef": "off"}

# Galveston Bay. Anything outside this is a typo, not a sample.
LAT_RANGE = (28.5, 30.2)
LON_RANGE = (-95.5, -94.0)

SHEET_RE = re.compile(r"^\s*(\d+)\s*\(([^)]+)\)\s*$")


def normalise(value) -> str:
    """Collapse the workbook's erratic whitespace so lookups are stable."""
    return re.sub(r"\s+", " ", str(value or "")).strip().lower()


def parse_sheet_title(title: str) -> tuple[int, str]:
    """'75 (GB20)' -> (75, 'GB20')"""
    match = SHEET_RE.match(title)
    if not match:
        raise ValueError(f"Unexpected worksheet name {title!r}; expected '<app#> (<site code>)'.")
    return int(match.group(1)), match.group(2).strip()


def read_acreage(ws) -> dict[str, float | None]:
    """The H/I summary block: on-reef and off-reef acreage for the site."""
    acres: dict[str, float | None] = {"on": None, "off": None}
    for row in ws.iter_rows(min_row=2, max_row=6, min_col=8, max_col=9, values_only=True):
        label, value = (row + (None, None))[:2]
        key = REEF_TYPES.get(normalise(label))
        if key and isinstance(value, (int, float)):
            acres[key] = float(value)
    return acres


def read_points(ws, app_no: int) -> list[dict]:
    points: list[dict] = []
    seen: set[int] = set()

    for index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        cells = (list(row) + [None] * 6)[:6]
        sheet_app, point_no, lat, lon, reef, datum = cells

        # Rows that only carry the H/I summary have an empty A column.
        if sheet_app is None and point_no is None:
            continue

        if int(sheet_app) != app_no:
            raise ValueError(
                f"{ws.title} row {index}: App# {sheet_app} does not match the sheet's {app_no}."
            )

        reef_type = REEF_TYPES.get(normalise(reef))
        if reef_type is None:
            raise ValueError(f"{ws.title} row {index}: unrecognised Reef value {reef!r}.")

        if not isinstance(lat, (int, float)) or not isinstance(lon, (int, float)):
            raise ValueError(f"{ws.title} row {index}: non-numeric coordinate {lat!r}, {lon!r}.")

        if not (LAT_RANGE[0] <= lat <= LAT_RANGE[1] and LON_RANGE[0] <= lon <= LON_RANGE[1]):
            raise ValueError(
                f"{ws.title} row {index}: point {point_no} at {lat}, {lon} is outside Galveston Bay."
            )

        # WGS84 is what the app and MapLibre assume; a sheet in anything else
        # would be silently wrong by tens of metres.
        if "wgs" not in normalise(datum):
            raise ValueError(f"{ws.title} row {index}: unexpected datum {datum!r}, expected WGS 1984.")

        point_no = int(point_no)
        if point_no in seen:
            raise ValueError(f"{ws.title}: point {point_no} appears twice.")
        seen.add(point_no)

        points.append(
            {
                "app_no": app_no,
                "point_no": point_no,
                "lat": round(float(lat), 6),
                "lon": round(float(lon), 6),
                "reef_type": reef_type,
            }
        )

    return sorted(points, key=lambda p: p["point_no"])


def sql_literal(value) -> str:
    if value is None:
        return "null"
    if isinstance(value, (int, float)):
        return repr(value)
    return "'" + str(value).replace("'", "''") + "'"


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)

    sites: list[dict] = []
    points: list[dict] = []

    for ws in wb.worksheets:
        app_no, site_code = parse_sheet_title(ws.title)
        sheet_points = read_points(ws, app_no)
        acres = read_acreage(ws)

        sites.append(
            {
                "app_no": app_no,
                "site_code": site_code,
                "on_reef_acres": acres["on"],
                "off_reef_acres": acres["off"],
                "on_reef_points": sum(1 for p in sheet_points if p["reef_type"] == "on"),
                "off_reef_points": sum(1 for p in sheet_points if p["reef_type"] == "off"),
            }
        )
        points.extend(sheet_points)

    sites.sort(key=lambda s: s["app_no"])
    points.sort(key=lambda p: (p["app_no"], p["point_no"]))

    JSON_OUT.write_text(
        json.dumps({"sites": sites, "points": points}, indent=2) + "\n", encoding="utf-8"
    )

    lines = [
        "-- Assigned ground samples, generated by scripts/extract_survey_points.py.",
        "-- Do not edit by hand: re-run the script when TPWD sends more points.",
        "-- Re-runnable -- on conflict the assignment is refreshed, and nothing here",
        "-- touches survey_samples, so re-seeding never disturbs collected data.",
        "",
    ]

    lines.append("insert into public.survey_sites (app_no, site_code, on_reef_acres, off_reef_acres) values")
    lines.append(
        ",\n".join(
            "  ({}, {}, {}, {})".format(
                s["app_no"], sql_literal(s["site_code"]),
                sql_literal(s["on_reef_acres"]), sql_literal(s["off_reef_acres"]),
            )
            for s in sites
        )
        + "\non conflict (app_no) do update set"
        + "\n  site_code = excluded.site_code,"
        + "\n  on_reef_acres = excluded.on_reef_acres,"
        + "\n  off_reef_acres = excluded.off_reef_acres;"
    )
    lines.append("")

    lines.append("insert into public.survey_points (app_no, point_no, lat, lon, reef_type) values")
    lines.append(
        ",\n".join(
            "  ({}, {}, {}, {}, {})".format(
                p["app_no"], p["point_no"], p["lat"], p["lon"], sql_literal(p["reef_type"])
            )
            for p in points
        )
        + "\non conflict (app_no, point_no) do update set"
        + "\n  lat = excluded.lat,"
        + "\n  lon = excluded.lon,"
        + "\n  reef_type = excluded.reef_type;"
    )
    lines.append("")

    SQL_OUT.write_text("\n".join(lines), encoding="utf-8")

    on = sum(s["on_reef_points"] for s in sites)
    off = sum(s["off_reef_points"] for s in sites)
    print(f"{len(sites)} sites, {len(points)} points ({on} on-reef, {off} off-reef)")
    for s in sites:
        print(
            f"  {s['app_no']:>4} ({s['site_code']}): "
            f"{s['on_reef_points']:>3} on / {s['off_reef_points']:>3} off"
        )
    print(f"-> {JSON_OUT.relative_to(ROOT)}")
    print(f"-> {SQL_OUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
