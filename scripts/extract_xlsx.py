"""
Convert the COL workbooks -> data/applications.json + supabase/seed.sql

Base data comes from Justin_Johny_COL_Status.xlsx. "JW_JJ_Hanna mods.xlsx" then
overrides a handful of applications whose boundaries were redrawn to clear O&G
locations and TPWD restoration reefs; see apply_mods() for exactly what it
touches.

The workbook stores each polygon vertex as a pair of `Latitude N` / `Longitude N`
columns (N = 1..10) in the order TPWD supplied them. This flattens that wide
layout into GeoJSON-style ring coordinates ([lon, lat], first vertex repeated
to close the ring) and emits both a JSON seed for the app and idempotent SQL
for Supabase.

Only the `GIS Upload` sheet is read -- the `Justin` and `Johny` tabs are
group-filtered copies of the same rows.
"""

from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "Justin_Johny_COL_Status.xlsx"
JSON_OUT = ROOT / "data" / "applications.json"
SQL_OUT = ROOT / "supabase" / "seed.sql"
MODS_XLSX = ROOT / "JW_JJ_Hanna mods.xlsx"
MODS_SQL_OUT = ROOT / "supabase" / "apply_mods.sql"

MAX_VERTICES = 10
VALID_STATUS = {"Accept", "Modify", "Decline"}


def signed_area(ring: list[list[float]]) -> float:
    """Shoelace on [lon, lat] pairs. Positive = counter-clockwise."""
    total = 0.0
    for i in range(len(ring) - 1):
        x1, y1 = ring[i]
        x2, y2 = ring[i + 1]
        total += x1 * y2 - x2 * y1
    return total / 2.0


def segments_cross(a1, a2, b1, b2) -> bool:
    """Proper segment intersection test (shared endpoints don't count)."""

    def orient(p, q, r):
        v = (q[1] - p[1]) * (r[0] - q[0]) - (q[0] - p[0]) * (r[1] - q[1])
        if abs(v) < 1e-12:
            return 0
        return 1 if v > 0 else 2

    o1, o2 = orient(a1, a2, b1), orient(a1, a2, b2)
    o3, o4 = orient(b1, b2, a1), orient(b1, b2, a2)
    return o1 != o2 and o3 != o4


def self_intersects(ring: list[list[float]]) -> bool:
    edges = [(ring[i], ring[i + 1]) for i in range(len(ring) - 1)]
    n = len(edges)
    for i in range(n):
        for j in range(i + 1, n):
            # skip adjacent edges (and the wrap-around pair) -- they share a vertex
            if j == i + 1 or (i == 0 and j == n - 1):
                continue
            if segments_cross(*edges[i], *edges[j]):
                return True
    return False


def slugify(value: str) -> str:
    return "".join(c.lower() if c.isalnum() else "-" for c in value).strip("-")


def parse_coordinate_blob(text: str) -> list[list[float]]:
    """
    The mods workbook keeps a whole ring in one cell as free text:
        "29.47568, -94.70283  29.47257, -94.69783  ..."
    Separators are inconsistent (some pairs lack the comma), so pull the
    numbers out in order and read them as alternating lat, lon. Returns
    [lon, lat] pairs to match GeoJSON.
    """
    values = [float(v) for v in re.findall(r"-?\d+\.\d+", text or "")]
    if len(values) % 2:
        raise ValueError(f"odd number of coordinate values ({len(values)})")
    return [[values[i + 1], values[i]] for i in range(0, len(values), 2)]


def build_ring(points: list[list[float]]) -> list[list[float]]:
    """Dedupe consecutive repeats, close the ring, force counter-clockwise."""
    deduped = [points[0]]
    for point in points[1:]:
        if point != deduped[-1]:
            deduped.append(point)
    closed = deduped + [deduped[0]]
    if signed_area(closed) < 0:
        closed = list(reversed(closed))
    return closed


def apply_mods(records: list[dict], notes: list[str], problems: list[str]) -> list[int]:
    """
    Overlay "JW_JJ_Hanna mods.xlsx" onto the base records.

    For every row carrying an Application#, the `Modified Coordinates` cell
    replaces the geometry and `New acreage` replaces the acreage. Status is
    forced to Modify: every application in this workbook had its boundary
    redrawn, which is a modification regardless of what the sheet's several
    disagreeing Status/Reevaluation columns say.

    `Corner Coordinates` holds the pre-modification ring. It is not imported --
    it is only checked against what we already have, so a mismatch means the
    two workbooks have drifted apart and the override is unsafe.
    """
    if not MODS_XLSX.exists():
        return []

    by_id = {r["id"]: r for r in records}
    wb = openpyxl.load_workbook(MODS_XLSX, data_only=True)
    changed: list[int] = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        header = list(rows[0])
        for raw in rows[1:]:
            row = dict(zip(header, raw))
            app_no = row.get("Application#")
            if app_no is None:
                continue
            app_no = int(app_no)

            record = by_id.get(app_no)
            if record is None:
                problems.append(f"mods: app {app_no} is not in the base workbook")
                continue

            try:
                original = parse_coordinate_blob(row.get("Corner Coordinates"))
                modified = parse_coordinate_blob(row.get("Modified Coordinates"))
            except ValueError as exc:
                problems.append(f"mods: app {app_no}: {exc}")
                continue

            if len(modified) < 3:
                problems.append(f"mods: app {app_no}: only {len(modified)} modified vertices")
                continue

            # Guard: the sheet's "original" ring must be what we already hold,
            # otherwise the two workbooks have drifted and overwriting is unsafe.
            # If it already matches the modified ring the overlay simply ran
            # before, which keeps this script re-runnable.
            def fingerprint(points):
                return sorted((round(x, 5), round(y, 5)) for x, y in points)

            current = record["geometry"]["coordinates"][0][:-1]
            if fingerprint(current) == fingerprint(original):
                pass
            elif fingerprint(current) == fingerprint(build_ring(modified)[:-1]):
                notes.append(f"mods: app {app_no} already carried the modified boundary")
            else:
                problems.append(
                    f"mods: app {app_no}: Corner Coordinates match neither the current "
                    f"geometry nor the modification -- refusing to overwrite"
                )
                continue

            closed = build_ring(modified)
            if self_intersects(closed):
                problems.append(f"mods: app {app_no}: modified ring self-intersects")
                continue

            before_acres = record["acreage"]
            before_status = record["status"]
            record["geometry"] = {"type": "Polygon", "coordinates": [closed]}
            if row.get("New acreage") is not None:
                record["acreage"] = float(row["New acreage"])
            record["status"] = "Modify"

            notes.append(
                f"mods: app {app_no} {sheet.title}: {len(current)}->{len(closed) - 1} vertices, "
                f"{before_acres}->{record['acreage']} ac, status {before_status}->Modify"
            )
            changed.append(app_no)

    return sorted(changed)


def main() -> None:
    records: list[dict] = []
    problems: list[str] = []
    notes: list[str] = []

    if XLSX.exists():
        rows = list(openpyxl.load_workbook(XLSX, data_only=True)["GIS Upload"]
                    .iter_rows(values_only=True))
        header = list(rows[0])
        data_rows = rows[1:]
    else:
        # The base workbook is optional. data/applications.json is the
        # committed extract of it, so the mods overlay still has something
        # to apply to when only the mods workbook is on hand.
        notes.append(f"{XLSX.name} not found -- reusing {JSON_OUT.name} as the base")
        records = json.loads(JSON_OUT.read_text(encoding="utf-8"))
        header, data_rows = [], []

    for raw in data_rows:
        row = dict(zip(header, raw))
        app_no = row.get("TPWD Application #")
        if app_no is None:
            continue

        ring: list[list[float]] = []
        for i in range(1, MAX_VERTICES + 1):
            lat = row.get(f"Latitude {i}")
            lon = row.get(f"Longitude {i}")
            if lat is None or lon is None:
                continue
            ring.append([float(lon), float(lat)])

        declared = row.get("Coordinate Count")
        if declared is not None and int(declared) != len(ring):
            problems.append(
                f"app {app_no}: Coordinate Count says {declared} but found {len(ring)} vertices"
            )

        # A few source rows repeat a vertex (app 107 lists vertex 3 twice). The
        # zero-length edge that creates makes the ring degenerate, so drop
        # consecutive duplicates before anything downstream sees the geometry.
        deduped = [ring[0]]
        for point in ring[1:]:
            if point != deduped[-1]:
                deduped.append(point)
        if len(deduped) != len(ring):
            notes.append(
                f"app {app_no}: dropped {len(ring) - len(deduped)} duplicate vertex/vertices"
            )
        ring = deduped

        if len(ring) < 3:
            problems.append(f"app {app_no}: only {len(ring)} vertices -- cannot form a polygon")
            continue

        status = row.get("Status")
        if status not in VALID_STATUS:
            problems.append(f"app {app_no}: unexpected status {status!r}")

        closed = ring + [ring[0]]
        # GeoJSON RFC 7946 wants exterior rings counter-clockwise.
        if signed_area(closed) < 0:
            closed = list(reversed(closed))
        if self_intersects(closed):
            problems.append(f"app {app_no}: ring self-intersects (bowtie) -- renders oddly")

        # Shaped exactly like a row of public.col_applications, so the offline
        # seed and the database return identical objects to the app.
        records.append(
            {
                "id": int(app_no),
                "group_name": row.get("Group"),
                "status": status,
                "applicant": row.get("Applicant Name"),
                "bay_system": row.get("Bay System"),
                "acreage": float(row["Acreage"]) if row.get("Acreage") is not None else None,
                "geometry": {"type": "Polygon", "coordinates": [closed]},
            }
        )

    modified_ids = apply_mods(records, notes, problems)

    records.sort(key=lambda r: r["id"])

    JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
    SQL_OUT.parent.mkdir(parents=True, exist_ok=True)
    JSON_OUT.write_text(json.dumps(records, indent=2), encoding="utf-8")

    lines = [
        "-- Generated by scripts/extract_xlsx.py -- do not edit by hand.",
        "-- Re-runnable: existing rows keep their current status, everything else is refreshed.",
        "",
    ]
    for r in records:
        geom = json.dumps(r["geometry"]).replace("'", "''")
        applicant = r["applicant"].replace("'", "''")
        lines.append(
            "insert into public.col_applications "
            "(id, group_name, status, applicant, bay_system, acreage, geometry) values "
            f"({r['id']}, '{r['group_name']}', '{r['status']}', '{applicant}', "
            f"'{r['bay_system']}', {r['acreage']}, '{geom}'::jsonb)"
            "\n  on conflict (id) do update set"
            "\n    group_name = excluded.group_name,"
            "\n    applicant  = excluded.applicant,"
            "\n    bay_system = excluded.bay_system,"
            "\n    acreage    = excluded.acreage,"
            "\n    geometry   = excluded.geometry;"
        )
    SQL_OUT.write_text("\n".join(lines) + "\n", encoding="utf-8")

    # seed.sql deliberately never touches `status`, so re-seeding cannot revive
    # a decision someone already made in the app. The mods DO change status, so
    # they get their own explicit statements to run against a live database.
    if modified_ids:
        mod_lines = [
            "-- Generated by scripts/extract_xlsx.py from 'JW_JJ_Hanna mods.xlsx'.",
            "-- Boundary modifications: new geometry, new acreage, status -> Modify.",
            "-- Safe to re-run.",
            "",
        ]
        for app_no in modified_ids:
            r = next(x for x in records if x["id"] == app_no)
            geom = json.dumps(r["geometry"]).replace("'", "''")
            mod_lines.append("update public.col_applications set")
            mod_lines.append(f"  status   = '{r['status']}',")
            mod_lines.append(f"  acreage  = {r['acreage']},")
            mod_lines.append(f"  geometry = '{geom}'::jsonb")
            mod_lines.append(f"where id = {app_no};")
            mod_lines.append("")
        MODS_SQL_OUT.write_text("\n".join(mod_lines) + "\n", encoding="utf-8")

    print(f"wrote {len(records)} records")
    print(f"  {JSON_OUT.relative_to(ROOT)}")
    print(f"  {SQL_OUT.relative_to(ROOT)}")
    if modified_ids:
        print(f"  {MODS_SQL_OUT.relative_to(ROOT)}  ({len(modified_ids)} boundary modifications)")
    owners = sorted({r["applicant"] for r in records})
    print(f"owners: {len(owners)}, bays: {sorted({r['bay_system'] for r in records})}")
    if notes:
        print(f"\n{len(notes)} cleanup(s) applied:")
        for n in notes:
            print("  ~", n)
    if problems:
        print(f"\n{len(problems)} data issue(s):")
        for p in problems:
            print("  !", p)
    else:
        print("\nno geometry or status issues found")


if __name__ == "__main__":
    main()
